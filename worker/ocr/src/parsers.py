from __future__ import annotations

import csv
import os
import re
import shutil
import subprocess
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from defusedxml import ElementTree as SafeET
from PIL import Image, ImageOps, ImageSequence, UnidentifiedImageError
from pillow_heif import register_heif_opener

try:
    from pillow_heif import register_avif_opener
except ImportError:  # pragma: no cover - guarded for older pinned wheels
    register_avif_opener = None

from pypdf import PdfReader
from pypdf.errors import PdfReadError

from .contract import HEBREW_LINE_ORDER, HEBREW_VISUAL_ORDER, validate_extraction
from .errors import ProcessingError
from .limits import DEFAULT_LIMITS, ExtractionLimits
from .mime import mime_matches, sniff_mime, inspect_zip
from .ocr import DisabledOcrAdapter, OcrAdapter, PageImage
from .second_pass import improve as improve_failed_pages
from .tempfiles import job_temp_dir


register_heif_opener()
if register_avif_opener is not None:
    register_avif_opener()

FULL_BBOX = [0.0, 0.0, 1.0, 1.0]
IMAGE_MIME = {"image/jpeg", "image/png", "image/webp", "image/heic", "image/heif", "image/gif", "image/avif"}


# A Hebrew word, allowing the quote marks that appear inside abbreviations such as בע"מ but never
# swallowing a quote that belongs to adjacent Latin text.
HEBREW_RUN = re.compile("[֐-׿]+(?:[\"'׳״][֐-׿]+)*")
# ך ם ן ף ץ -- in correct Hebrew these appear only as the last letter of a word. A generator that
# laid the page out in visual order puts them first instead, which is a lexicon-free tell.
HEBREW_FINALS = "ךםןףץ"


def _hebrew_order_evidence(text: str) -> tuple[int, int, int]:
    """The three numbers `_hebrew_is_reversed` decides on, returned instead of thrown away.

    `(judged_words, final_letter_first, final_letter_last)`. The decision is one comparison over
    the last two, so keeping them keeps the REASON: a reader of a stored extraction can see what
    the detector saw, not merely that it fired. Words shorter than two letters cannot carry the
    signal -- a single ם is both the first and the last letter -- and are counted by neither.
    """
    words = starts = ends = 0
    for word in HEBREW_RUN.findall(text):
        if len(word) < 2:
            continue
        words += 1
        if word[0] in HEBREW_FINALS:
            starts += 1
        if word[-1] in HEBREW_FINALS:
            ends += 1
    return words, starts, ends


def _hebrew_is_reversed(text: str) -> bool:
    """Detects a text layer stored in visual rather than logical order.

    Measured on a real Israeli supplier receipt: 13 Hebrew words began with a final letter and
    none ended with one. The same check over spreadsheet text and over image OCR returned 0 vs
    1701 and 0 vs 75, so the separation is unambiguous in both directions.
    """
    _, starts, ends = _hebrew_order_evidence(text)
    return starts > ends


def _reverse_hebrew_runs(text: str) -> str:
    """Restores logical order WITHIN each Hebrew word. Word order is not touched.

    This is half of the repair, and for years it was the whole of it. See
    `_restore_line_order` below for the half that was missing and what it cost.
    """
    return HEBREW_RUN.sub(lambda match: match.group(0)[::-1], text)


# A run a bidirectional layout keeps together and left-to-right on the page: Latin letters,
# digits, and the separators that live INSIDE a number or a code -- `1,392.00`, `30*30`, `100/1`.
# Their internal order is already logical in a visual-order layer, because the printed page has to
# be readable: a generator that emitted `1/001` would have printed `1/001`, and no supplier ships
# a price list like that.
#
# THE SPACE INSIDE THE CLASS IS LOAD-BEARING, and it was found by measurement rather than by
# reading. Without it `מפיות דמוי בד לבן PREMIUM NAPKINS` -- one of the bilingual names W0-G8
# explicitly reported as UNDAMAGED -- came back as `... NAPKINS PREMIUM`, because two Latin words
# separated by a space were treated as two runs and swapped. A bidirectional layout keeps a whole
# left-to-right SEQUENCE together, spaces included. The alternation only absorbs a space when
# another Latin or digit token follows it, so `100 מ״ל` still breaks after `100`.
LTR_RUN = re.compile(r"[A-Za-z0-9]+(?:(?:[.,:/*+\-]|[ ])[A-Za-z0-9]+)*[%°]?")
# Brackets are mirrored when a run order flips. `(` at the end of a right-to-left line is drawn as
# `)` at its left edge, and an extractor that reads glyphs reads the mirrored character.
MIRRORED = {"(": ")", ")": "(", "[": "]", "]": "[", "{": "}", "}": "{", "<": ">", ">": "<"}
BRACKET_PAIRS = (("(", ")"), ("[", "]"), ("{", "}"))

# How a Hebrew document enumerates: a single letter or a small number, a closer, then the item.
# `א) מוצר`, `1) קמח`, `ב. שמן` -- and the first two put a closer after one character with content
# after it, which is precisely the shape the inversion scan treats as a mirrored bracket. Counting
# them made every itemised price list look reversed, so the marker is stripped before a line is
# judged. Deliberately narrow: ONE character before the closer. `12) x` is a marker; `(100 יח` is
# not, and `100 יח)` is the stray-closer case that has its own rule.
#
# The trailing separator is a LOOKAHEAD, not a consumed space: OCR drops the space after the
# marker often enough that `א)קמח` is ordinary, and requiring `\s+` let exactly that spelling
# score as inversion evidence while `א) קמח` did not.
#
# AND THE LOOKAHEAD ADMITS ANY CONTENT, not a chosen alphabet. Naming the characters that may
# follow — whitespace, a digit, a Hebrew letter — left `א)₪12 קמח` and `א)FLOUR` outside it, so a
# price list that opens an item with a currency sign or a Latin product name still read as
# reversed. What makes this a marker is its SHAPE at the start of the line, not what comes after
# it. A line that opens with the closer itself never matches, because the pattern requires one or
# two ordinary characters before it.
LIST_MARKER = re.compile(r"^[0-9֐-׿]{1,2}\s*[).\]](?=.)")


def _line_order_evidence(text: str) -> tuple[int, int, int, int, int, int]:
    """`(lines_judged, leading_closer, closer_before_opener, unbalanced, strong, evidence)`.

    THE SAME THREE SIGNALS THE PRODUCT ALREADY COUNTS. `src/lib/productDisplayName.ts` and
    `scripts/report-product-name-health.ts` carry this detector on the client side, pinned against
    each other by `productDisplayName.spec.ts`, and W0-G8 measured production with it: 6 names
    began with a closing bracket and 27 carried a close with no open. This is that detector, in
    the one place that could have prevented the damage instead of reporting it.

    Brackets are the cheap, reliable tell. In logical order an opener precedes its closer; a line
    stored in visual order has them the other way round. `unbalanced` is recorded but does NOT
    decide, because OCR loses a bracket often enough that it would fire on undamaged text.

    A STRAY CLOSER AT THE END OF A LINE IS NOT EVIDENCE, and counting it as evidence was a real
    defect. `קמח לבן 5 ק"ג)` -- an ordinary product name with one dropped opener -- used to score
    `closer_before_opener = 1` and, because the decision is taken for the WHOLE DOCUMENT, that one
    line was enough to reverse the letters of every line on every page. The discriminator is not
    "is the line balanced", because a line whose opener OCR dropped is unbalanced either way. It is
    WHERE the unmatched closer sits:

      `שקיות אשפה 60*80 )100 יח`  the closer comes BEFORE the `100 יח` it would enclose. No
                                  logical order puts it there; it is a mirrored `(` that travelled
                                  with the reversed run. Evidence.
      `קמח לבן 5 ק"ג)`            the closer ENDS the line, which is exactly where a closer belongs.
                                  The bracket that went missing is the opener. Not evidence.

    So an unmatched closer counts only when real content follows it on the same line. A closer that
    is matched but out of order -- `)…(` on a balanced line -- cannot happen in logical order at
    all, so it counts on its own and is reported as `strong` alongside a leading closer.

    `strong` and `evidence` are counts of LINES, not of tells, and they are what the document-level
    decision is allowed to read. One line can raise both `leading_closer` and
    `closer_before_opener`, so summing those two would have let a single line corroborate itself.
    """
    lines = judged = leading = inverted = unbalanced = strong = evidence = 0
    # Unclosed openers carried over from earlier lines, per pair. A bracket that OPENS on one line
    # and CLOSES on the next is ordinary logical text -- a sentence in parentheses that wrapped --
    # and judging each line as if it were a whole paragraph read the second line's leading closer
    # as a mirrored bracket. Measured: `(תנאים / ) המשך / א) מוצר` scored strong=1, evidence=2 and
    # inverted a valid document.
    pending = {closer: 0 for _, closer in BRACKET_PAIRS}
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        lines += 1
        if not HEBREW_RUN.search(line):
            # A line with no Hebrew carries no evidence about Hebrew layout either way, but its
            # brackets still open and close, so the running depth has to see it.
            for opener, closer in BRACKET_PAIRS:
                pending[closer] = max(0, pending[closer] + line.count(opener) - line.count(closer))
            continue
        judged += 1
        # A HEBREW LIST MARKER IS NOT A MIRRORED BRACKET. `א) מוצר` and `1) מוצר` are how a Hebrew
        # document enumerates, and they put a closer after a single character with content after
        # it -- which is exactly the shape the inversion scan below looks for. Counting them made
        # any itemised price list look reversed. The marker is removed before the line is judged,
        # and its closer is not allowed to satisfy the depth bookkeeping either.
        marker = LIST_MARKER.match(line)
        if marker:
            line = line[marker.end() :].strip()
            if not line or not HEBREW_RUN.search(line):
                judged -= 1
                continue
        # A leading closer is evidence ONLY when nothing earlier left an opener hanging.
        line_leading = any(
            line.startswith(close) and pending[close] == 0 for _, close in BRACKET_PAIRS
        )
        if line_leading:
            leading += 1
        for opener, closer in BRACKET_PAIRS:
            opens = line.count(opener)
            closes = line.count(closer)
            if opens != closes:
                unbalanced += 1
                break
        line_inverted = False
        line_strong = False
        for opener, closer in BRACKET_PAIRS:
            # START FROM WHAT EARLIER LINES LEFT OPEN. Resetting to zero on every line made a
            # chain of continuation lines read as balanced-and-inverted: each one closes the
            # previous line's opener and opens its own, which is ordinary wrapped prose. Carrying
            # the depth into the SCAN, and not only into the skip test above, is what makes the
            # cross-line bookkeeping actually decide anything.
            depth = pending[closer]
            position = -1
            for index, char in enumerate(line):
                if char == opener:
                    depth += 1
                elif char == closer:
                    if depth == 0:
                        position = index
                        break
                    depth -= 1
            if position < 0:
                continue
            # NO NET-COUNT SHORTCUT HERE. It used to stand where this comment does, and once the
            # scan above began at `pending[closer]` it became both redundant and wrong: the scan
            # has already spent the carried debt, so a `position` it returns is a closer nothing
            # could match. Counting again over the whole line let a LATER opener cancel that
            # finding — `) המשך )100 יח (` balances on the net and is still mirrored — and two
            # genuinely damaged lines went unrepaired because of it.
            balanced = line.count(opener) == line.count(closer)
            # Real content, not a trailing quote or full stop: `(abc).` with the opener dropped
            # reads as `abc).`, and the `.` must not be mistaken for the text a closer preceded.
            content_follows = any(char.isalnum() for char in line[position + 1 :])
            if balanced or content_follows:
                line_inverted = True
                if balanced:
                    line_strong = True
                break
        if line_inverted:
            inverted += 1
        if line_leading or line_strong:
            strong += 1
        if line_leading or line_inverted:
            evidence += 1
        # Carry this line's unclosed openers to the next one.
        for opener, closer in BRACKET_PAIRS:
            pending[closer] = max(0, pending[closer] + line.count(opener) - line.count(closer))
    return judged, leading, inverted, unbalanced, strong, evidence


def _restore_line_order(text: str) -> str:
    """Invert a visual-order LINE: the whole line, not each word separately.

    WHAT WENT WRONG WITHOUT THIS, MEASURED. `_reverse_hebrew_runs` turns each Hebrew word the
    right way round and leaves everything else exactly where it sits -- the self-check has said
    so in as many words for a long time, returning `"word_order": "not_repaired_by_design"`. For
    an invoice that is enough: a reader wants field values, not sentences. For a PRODUCT NAME the
    order IS the value, and the residue is exactly what W0-G8 counted in production: of 271
    catalogue names, 105 carry a bracket at the wrong end or a digit fused to the Hebrew letter
    that belonged after it -- `)ק"ג 5( קמח לבן` for `קמח לבן (5 ק"ג)`.

    THE TRANSFORM. Split the line into runs; emit them in reverse; reverse each Hebrew run
    internally; leave each left-to-right run alone; mirror each bracket. That is the Unicode
    bidirectional reordering for a right-to-left paragraph, and running it on visual order gives
    logical order back.

    IT IS ITS OWN INVERSE. The run sequence is reversed twice, each Hebrew run twice, each bracket
    mirrored twice -- and reversing a Hebrew run leaves it a Hebrew run, so the second pass finds
    the same run boundaries. The evidence contract depends on this: re-applying the named
    transform to the preserved `original_text` must reproduce the stored text, and the self-check
    asserts it rather than trusting the argument.
    """
    out: list[str] = []
    for line in text.split("\n"):
        if not HEBREW_RUN.search(line):
            # A line with no Hebrew was never laid out right-to-left, so there is nothing to
            # invert. Without this, `P18B product` came back as `product P18B`: a correction
            # inventing damage on a line that had none.
            out.append(line)
            continue
        runs: list[str] = []
        position = 0
        while position < len(line):
            match = HEBREW_RUN.match(line, position) or LTR_RUN.match(line, position)
            if match and match.end() > position:
                runs.append(match.group(0))
                position = match.end()
            else:
                runs.append(line[position])
                position += 1
        rebuilt = []
        for run in reversed(runs):
            if HEBREW_RUN.fullmatch(run):
                rebuilt.append(run[::-1])
            elif len(run) == 1:
                rebuilt.append(MIRRORED.get(run, run))
            else:
                rebuilt.append(run)
        out.append("".join(rebuilt))
    return "\n".join(out)


def _normalize_pdf_text_layer(
    native_text: dict[int, str]
) -> tuple[dict[int, str], list[dict[str, Any]]]:
    """Repairs a visual-order text layer AND returns the record of that decision beside it.

    THE POINT OF RETURNING TWO THINGS. This correction used to be a one-line in-place overwrite.
    It fixed the text and destroyed the question: afterwards nothing in the database could say
    whether a document had printed Hebrew logically or whether this worker had turned it around,
    so "what the supplier's document said" and "what this system decided it said" were one field
    with one value. Every downstream reading -- a supplier name, an invoice number, a product
    description a person later swears they never typed -- inherited that ambiguity.

    `original_text` is the EXACT string the detector judged: the pages joined with a single
    newline, before a character moved. That matters for verification rather than for display --
    `_reverse_hebrew_runs` is its own inverse over these runs, so applying it to `original_text`
    must reproduce the corrected pages joined the same way, and a test can say so instead of
    trusting it.

    Whole document, one decision, unchanged from before: a single page may not carry enough final
    letters to judge, and a per-page vote would repair some pages of an invoice and not others.

    TWO CORRECTIONS NOW, AND AT MOST ONE OF THEM FIRES.

      `hebrew_line_order`   -- the whole line was laid out backwards. Inverting the line already
                               turns every Hebrew word the right way round, so this SUPERSEDES the
                               word-level repair rather than running after it.
      `hebrew_visual_order` -- the words were stored backwards but their order on the line was
                               not. This is what shipped before, unchanged, and it is still the
                               right answer for every generator that behaves that way.

    Both are recorded either way, because "evaluated and left it alone" and "never looked" are
    different facts and the stored evidence has to keep them apart.

    WHY THE LINE DECISION IS NOT PER LINE. A price list has hundreds of lines and only some carry
    a bracket. A per-line vote would inverse the bracketed lines and leave the rest, producing a
    document half in one order and half in the other -- which is worse than either. One layer, one
    layout, one decision.
    """
    original = "\n".join(native_text.values())
    words, first_letter, last_letter = _hebrew_order_evidence(original)
    (
        lines,
        leading_closer,
        closer_first,
        unbalanced,
        strong_lines,
        evidence_lines,
    ) = _line_order_evidence(original)
    # ONE LINE IS NOT A DOCUMENT. This used to read `leading_closer + closer_first > 0`: a single
    # bracket anywhere in a hundred-page price list reversed the letters of every line on every
    # page, and a false positive here does not merely fail to repair -- `_restore_line_order`
    # rewrites the text, so it DESTROYS a name that was fine. The decision now needs a line whose
    # bracket order cannot occur in logical text at all (a leading closer, or a balanced pair the
    # wrong way round) AND a second line that independently carries the tell. `unbalanced` is
    # still measured and still does not decide.
    line_applied = strong_lines > 0 and evidence_lines >= 2
    # `_hebrew_is_reversed`'s rule, spelled here so the counts that go into the record and the
    # comparison that decides are read from one measurement rather than two passes of the same
    # regex. The self-check asserts the two agree on both fixtures, so this cannot drift into a
    # second, quieter definition of "reversed".
    word_applied = (first_letter > last_letter) and not line_applied
    line_record = {
        "id": HEBREW_LINE_ORDER,
        "applied": line_applied,
        "original_text": original if line_applied else None,
        "measurements": [
            {"name": "hebrew_lines", "value": lines},
            {"name": "leading_closer", "value": leading_closer},
            {"name": "closer_before_opener", "value": closer_first},
            {"name": "unbalanced_brackets", "value": unbalanced},
            # The two the decision actually reads, so a stored record can be re-judged without
            # re-deriving them from the text. Both count LINES.
            {"name": "impossible_order_lines", "value": strong_lines},
            {"name": "evidence_lines", "value": evidence_lines},
        ],
    }
    word_record = {
        "id": HEBREW_VISUAL_ORDER,
        "applied": word_applied,
        # Non-null exactly when something changed. An unapplied correction storing a second copy
        # of the untouched text would be a claim that a correction happened.
        "original_text": original if word_applied else None,
        "measurements": [
            {"name": "hebrew_words", "value": words},
            {"name": "final_letter_first", "value": first_letter},
            {"name": "final_letter_last", "value": last_letter},
        ],
    }
    records = [line_record, word_record]
    if line_applied:
        return {page: _restore_line_order(text) for page, text in native_text.items()}, records
    if word_applied:
        return {page: _reverse_hebrew_runs(text) for page, text in native_text.items()}, records
    return native_text, records


def _languages(text: str) -> list[str]:
    result: list[str] = []
    if re.search(r"[\u0590-\u05ff]", text):
        result.append("he")
    if re.search(r"[A-Za-z]", text):
        result.append("en")
    return result


def _contract(
    page_count: int,
    plain_text: str,
    *,
    blocks: list[dict[str, Any]] | None = None,
    tables: list[dict[str, Any]] | None = None,
    marks: list[dict[str, Any]] | None = None,
    partial: bool = False,
    normalizations: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    return {
        "schema_version": "1",
        "document": {
            "page_count": page_count,
            "detected_languages": _languages(plain_text),
            "plain_text": plain_text,
            "partial": partial,
        },
        "blocks": blocks or [],
        "tables": tables or [],
        "marks": marks or [],
        # Empty means "no corrector ran on this path", which is the truth for a spreadsheet, a
        # CSV or an HTML page: none of them has a text layer that can be laid out backwards. It
        # does NOT mean "nothing needed correcting" -- only an entry saying `applied: false` can
        # claim that, and only the PDF path produces one.
        "normalizations": normalizations or [],
    }


def _text_blocks(parts: list[str], page: int = 1, prefix: str = "text") -> list[dict[str, Any]]:
    return [
        {
            "id": f"{prefix}-p{page}-{index}",
            "page": page,
            "type": "text",
            "bbox": FULL_BBOX.copy(),
            "text": text,
            "confidence": None,
        }
        for index, text in enumerate(parts, start=1)
        if text
    ]


def _cell(value: Any) -> dict[str, Any]:
    if value is None:
        text = ""
    elif isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    return {"text": text, "bbox": None}


def _spreadsheet_contract(
    sheets: list[tuple[str, list[list[dict[str, Any]]]]], limits: ExtractionLimits
) -> dict[str, Any]:
    tables: list[dict[str, Any]] = []
    blocks: list[dict[str, Any]] = []
    texts: list[str] = []
    total_rows = 0
    for page, (name, rows) in enumerate(sheets, start=1):
        total_rows += len(rows)
        if total_rows > limits.max_spreadsheet_rows:
            raise ProcessingError("spreadsheet_row_limit", "Spreadsheet row limit exceeded")
        if any(len(row) > limits.max_spreadsheet_columns for row in rows):
            raise ProcessingError(
                "spreadsheet_column_limit", "Spreadsheet column limit exceeded"
            )
        row_text = ["\t".join(cell["text"] for cell in row) for row in rows]
        sheet_text = "\n".join(([name] if name else []) + row_text)
        texts.append(sheet_text)
        tables.append({"id": f"sheet-{page}", "page": page, "bbox": FULL_BBOX.copy(), "rows": rows})
        blocks.append(
            {
                "id": f"sheet-block-{page}",
                "page": page,
                "type": "table",
                "bbox": FULL_BBOX.copy(),
                "text": sheet_text,
                "confidence": None,
            }
        )
    plain_text = "\n\n".join(texts)
    if len(plain_text) > limits.max_text_chars:
        raise ProcessingError("text_length_limit", "Extracted text exceeds the text limit")
    return _contract(max(1, len(sheets)), plain_text, blocks=blocks, tables=tables)


def _parse_xlsx(path: Path, limits: ExtractionLimits) -> dict[str, Any]:
    inspect_zip(path, limits)
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise ProcessingError("corrupt_document", "Spreadsheet is corrupt") from exc
    sheets: list[tuple[str, list[list[dict[str, Any]]]]] = []
    total_rows = 0
    try:
        if len(workbook.worksheets) > limits.max_pdf_pages:
            raise ProcessingError("page_limit", "Spreadsheet contains too many sheets")
        for sheet in workbook.worksheets:
            if (sheet.max_column or 0) > limits.max_spreadsheet_columns:
                raise ProcessingError("spreadsheet_column_limit", "Spreadsheet contains too many columns")
            rows: list[list[dict[str, Any]]] = []
            for values in sheet.iter_rows(values_only=True):
                if len(values) > limits.max_spreadsheet_columns:
                    raise ProcessingError(
                        "spreadsheet_column_limit",
                        "Spreadsheet contains too many columns",
                    )
                if not any(value is not None for value in values):
                    continue
                total_rows += 1
                if total_rows > limits.max_spreadsheet_rows:
                    raise ProcessingError("spreadsheet_row_limit", "Spreadsheet row limit exceeded")
                rows.append([_cell(value) for value in values])
            sheets.append((sheet.title, rows))
    finally:
        workbook.close()
    return _spreadsheet_contract(sheets, limits)


def _parse_xls(path: Path, limits: ExtractionLimits) -> dict[str, Any]:
    try:
        workbook = xlrd.open_workbook(filename=str(path), on_demand=True)
    except (OSError, xlrd.XLRDError) as exc:
        raise ProcessingError("corrupt_document", "Legacy spreadsheet is corrupt") from exc
    sheets: list[tuple[str, list[list[dict[str, Any]]]]] = []
    total_rows = 0
    try:
        if workbook.nsheets > limits.max_pdf_pages:
            raise ProcessingError("page_limit", "Spreadsheet contains too many sheets")
        for sheet in workbook.sheets():
            if sheet.ncols > limits.max_spreadsheet_columns:
                raise ProcessingError("spreadsheet_column_limit", "Spreadsheet contains too many columns")
            rows: list[list[dict[str, Any]]] = []
            for row_index in range(sheet.nrows):
                values = [sheet.cell_value(row_index, column) for column in range(sheet.ncols)]
                if not any(value not in (None, "") for value in values):
                    continue
                total_rows += 1
                if total_rows > limits.max_spreadsheet_rows:
                    raise ProcessingError("spreadsheet_row_limit", "Spreadsheet row limit exceeded")
                rows.append([_cell(value) for value in values])
            sheets.append((sheet.name, rows))
    finally:
        workbook.release_resources()
    return _spreadsheet_contract(sheets, limits)


def _read_utf8(path: Path, limits: ExtractionLimits) -> str:
    try:
        text = path.read_text(encoding="utf-8", errors="strict")
    except UnicodeDecodeError as exc:
        raise ProcessingError("invalid_utf8", "Text source must be strict UTF-8") from exc
    except OSError as exc:
        raise ProcessingError("source_unavailable", "Source document is unavailable", retryable=True) from exc
    if len(text) > limits.max_text_chars:
        raise ProcessingError("text_length_limit", "Text source exceeds the text limit")
    return text


def _parse_csv(path: Path, limits: ExtractionLimits) -> dict[str, Any]:
    text = _read_utf8(path, limits)
    try:
        dialect = csv.Sniffer().sniff(text[:16_384], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows: list[list[dict[str, Any]]] = []
    try:
        for index, row in enumerate(csv.reader(text.splitlines(), dialect), start=1):
            if index > limits.max_spreadsheet_rows:
                raise ProcessingError("spreadsheet_row_limit", "CSV row limit exceeded")
            if len(row) > limits.max_spreadsheet_columns:
                raise ProcessingError("spreadsheet_column_limit", "CSV contains too many columns")
            rows.append([_cell(value) for value in row])
    except csv.Error as exc:
        raise ProcessingError("corrupt_document", "CSV is malformed") from exc
    return _spreadsheet_contract([("", rows)], limits)


# Elements that hold part of the document somewhere other than in this file's text. `_parse_html`
# has no image, plugin or frame handling anywhere in it -- it reads markup and nothing else -- so
# whatever these point at is content the extraction demonstrably did not capture.
# script/style/noscript are excluded on purpose below and are program text, not document content,
# so they are deliberately absent from this set.
EMBEDDED_CONTENT_TAGS = {"img", "object", "embed", "iframe", "frame"}


class _DocumentHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.ignored = 0
        self.embedded = 0
        self.text: list[str] = []
        self.tables: list[list[list[str]]] = []
        self.table: list[list[str]] | None = None
        self.row: list[str] | None = None
        self.cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        del attrs
        # Counted before the chain below, and separately from it: `<img>` is a void element, so
        # `handle_startendtag` routes both `<img>` and `<img/>` through here.
        if not self.ignored and tag in EMBEDDED_CONTENT_TAGS:
            self.embedded += 1
        if tag in {"script", "style", "noscript"}:
            self.ignored += 1
        elif not self.ignored and tag == "table":
            self.table = []
        elif not self.ignored and tag == "tr" and self.table is not None:
            self.row = []
        elif not self.ignored and tag in {"td", "th"} and self.row is not None:
            self.cell = []

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript"} and self.ignored:
            self.ignored -= 1
        elif tag in {"td", "th"} and self.cell is not None and self.row is not None:
            self.row.append(" ".join(self.cell).strip())
            self.cell = None
        elif tag == "tr" and self.row is not None and self.table is not None:
            self.table.append(self.row)
            self.row = None
        elif tag == "table" and self.table is not None:
            self.tables.append(self.table)
            self.table = None

    def handle_data(self, data: str) -> None:
        if self.ignored:
            return
        text = " ".join(data.split())
        if not text:
            return
        self.text.append(text)
        if self.cell is not None:
            self.cell.append(text)


def _parse_html(path: Path, limits: ExtractionLimits) -> dict[str, Any]:
    text = _read_utf8(path, limits)
    parser = _DocumentHtmlParser()
    try:
        parser.feed(text)
        parser.close()
    except Exception as exc:
        raise ProcessingError("corrupt_document", "HTML is malformed") from exc
    plain_text = "\n".join(parser.text)
    tables: list[dict[str, Any]] = []
    total_rows = 0
    for index, raw_rows in enumerate(parser.tables, start=1):
        total_rows += len(raw_rows)
        if total_rows > limits.max_spreadsheet_rows:
            raise ProcessingError("spreadsheet_row_limit", "HTML table row limit exceeded")
        if any(len(row) > limits.max_spreadsheet_columns for row in raw_rows):
            raise ProcessingError(
                "spreadsheet_column_limit", "HTML table contains too many columns"
            )
        tables.append(
            {
                "id": f"html-table-{index}",
                "page": 1,
                "bbox": FULL_BBOX.copy(),
                "rows": [[_cell(value) for value in row] for row in raw_rows],
            }
        )
    # `partial=True` used to be unconditional here and meant nothing. What this parser can and
    # cannot see is decidable: `_read_utf8` refuses an oversized source rather than truncating it,
    # and every text node outside script/style/noscript is collected, so the markup itself is read
    # in full. The one thing that escapes is content that is not text in this file -- the bytes
    # behind an <img>, <object>, <embed> or <iframe>, which nothing here fetches or decodes. A
    # scanned invoice mailed as an HTML wrapper around one <img> is exactly that case, and it is
    # the case worth reporting.
    return _contract(
        1,
        plain_text,
        blocks=_text_blocks(parser.text, prefix="html"),
        tables=tables,
        partial=parser.embedded > 0,
    )


WORDPROCESSING_NAMESPACE = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

# Parts that print as part of the page but are NOT `word/document.xml`, which is the only part
# `_parse_docx` opens. Comments are excluded on purpose: they are review annotations that do not
# render with the document, so leaving them unread is not a coverage gap.
DOCX_PRINTED_TEXT_PARTS = re.compile(r"^word/(header\d*|footer\d*|footnotes|endnotes)\.xml$")


def _docx_ancillary_text(archive: zipfile.ZipFile) -> tuple[list[str], list[str]]:
    """Paragraph text from the printed parts that are not `word/document.xml`.

    Headers, footers, footnotes and endnotes print with the page but live in their own archive
    parts, and `_parse_docx` opens exactly one part by name. Reading them here is what keeps a
    supplier's letterhead -- company name, VAT number, address -- from being silently dropped,
    and it is also what stops a page-number footer from marking every Word document partial.
    Excluding them from the flag instead would have made the flag lie in the other direction.

    Placement is deliberately NOT reconstructed. Which header belongs to which page is a function
    of section properties and pagination that nothing here renders, so no `r:id` relationship is
    resolved and no interleaving is attempted: the text is appended in a stable part-name order.
    That is not a new liberty -- `_parse_docx` already appends all table text after all
    paragraphs, so "after the body" is this parser's existing convention for content whose
    position it does not model.

    `word/comments.xml` is excluded on purpose: review annotations do not render with the
    document, so they are not part of it. Returns the extracted paragraphs, and the names of any
    parts that could not be opened at all. `inspect_zip` has already bounded this archive's entry
    count, decompressed size and compression ratio.
    """
    paragraphs: list[str] = []
    unreadable: list[str] = []
    for name in sorted(archive.namelist()):
        if not DOCX_PRINTED_TEXT_PARTS.match(name):
            continue
        try:
            with archive.open(name) as part:
                root = SafeET.parse(part).getroot()
        except (OSError, KeyError, zipfile.BadZipFile, SafeET.ParseError):
            # A part this parser cannot open is a part it certainly did not extract, which is a
            # real coverage gap. Recording it is honest; raising would throw away a document whose
            # body read perfectly because its footer is malformed.
            unreadable.append(name)
            continue
        for paragraph in root.iter(f"{WORDPROCESSING_NAMESPACE}p"):
            # `iter` recurses, so a paragraph inside a header table is picked up too. Footnote
            # separators are paragraphs with no `w:t` text and fall out here.
            text = "".join(
                node.text or "" for node in paragraph.iter(f"{WORDPROCESSING_NAMESPACE}t")
            ).strip()
            if text:
                paragraphs.append(text)
    return paragraphs, unreadable


def _parse_docx(path: Path, limits: ExtractionLimits) -> dict[str, Any]:
    inspect_zip(path, limits)
    try:
        with zipfile.ZipFile(path) as archive:
            with archive.open("word/document.xml") as document_xml:
                root = SafeET.parse(document_xml).getroot()
            ancillary, unreadable_parts = _docx_ancillary_text(archive)
    except (OSError, KeyError, zipfile.BadZipFile, SafeET.ParseError) as exc:
        raise ProcessingError("corrupt_document", "Word document is corrupt") from exc
    namespace = WORDPROCESSING_NAMESPACE
    body = root.find(f"{namespace}body")
    if body is None:
        raise ProcessingError("corrupt_document", "Word document body is missing")
    paragraphs: list[str] = []
    tables: list[dict[str, Any]] = []
    total_rows = 0
    skipped_body_text = False
    for child in body:
        if child.tag == f"{namespace}p":
            text = "".join(node.text or "" for node in child.iter(f"{namespace}t")).strip()
            if text:
                paragraphs.append(text)
        elif child.tag == f"{namespace}tbl":
            rows: list[list[dict[str, Any]]] = []
            for row in child.findall(f"{namespace}tr"):
                total_rows += 1
                if total_rows > limits.max_spreadsheet_rows:
                    raise ProcessingError("spreadsheet_row_limit", "Word table row limit exceeded")
                raw_cells = row.findall(f"{namespace}tc")
                if len(raw_cells) > limits.max_spreadsheet_columns:
                    raise ProcessingError(
                        "spreadsheet_column_limit", "Word table contains too many columns"
                    )
                cells = []
                for cell in raw_cells:
                    cells.append(_cell("".join(node.text or "" for node in cell.iter(f"{namespace}t")).strip()))
                rows.append(cells)
            tables.append(
                {"id": f"docx-table-{len(tables) + 1}", "page": 1, "bbox": FULL_BBOX.copy(), "rows": rows}
            )
        elif any((node.text or "").strip() for node in child.iter(f"{namespace}t")):
            # The two branches above are the whole of this loop: a body child that is neither a
            # paragraph nor a table -- a content control (`w:sdt`), say -- is dropped on the floor.
            # Dropping one that carries text is a coverage gap; `w:sectPr` and bookmark markers
            # carry none and are therefore never counted here.
            skipped_body_text = True
    table_text = ["\n".join("\t".join(cell["text"] for cell in row) for row in table["rows"]) for table in tables]
    plain_text = "\n\n".join(paragraphs + table_text + ancillary)
    if len(plain_text) > limits.max_text_chars:
        raise ProcessingError("text_length_limit", "Extracted text exceeds the text limit")
    # `partial=True` used to be unconditional here too, and said nothing about any particular file.
    # What is left after the headers and footers are actually READ is a short, honest list: a body
    # child the loop above skipped while carrying text, and a printed part that would not open. An
    # ordinary supplier document -- paragraphs, a table, a letterhead and a page number -- now
    # reports False, and everything on it appears in `plain_text`.
    return _contract(
        1,
        plain_text,
        blocks=_text_blocks(paragraphs + ancillary, prefix="docx"),
        tables=tables,
        partial=skipped_body_text or bool(unreadable_parts),
    )


def _convert_to_docx(path: Path, limits: ExtractionLimits, source_suffix: str) -> Path:
    with job_temp_dir(path.parent) as conversion:
        profile = conversion / "profile"
        profile_user = profile / "user"
        profile_user.mkdir(parents=True)
        (profile_user / "registrymodifications.xcu").write_text(
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<oor:items xmlns:oor="http://openoffice.org/2001/registry">'
            '<item oor:path="/org.openoffice.Office.Common/Security/Scripting">'
            '<prop oor:name="MacroSecurityLevel" oor:op="fuse"><value>3</value></prop>'
            '</item></oor:items>',
            encoding="utf-8",
        )
        output = conversion / "output"
        output.mkdir()
        typed_source = conversion / f"source{source_suffix}"
        shutil.copyfile(path, typed_source)
        env = os.environ.copy()
        env.update(
            {
                "TMPDIR": str(conversion),
                "XDG_CACHE_HOME": str(conversion / "cache"),
                "XDG_CONFIG_HOME": str(conversion / "config"),
                "SAL_USE_VCLPLUGIN": "svp",
            }
        )
        command = [
            "libreoffice",
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            "--nofirststartwizard",
            f"-env:UserInstallation={profile.as_uri()}",
            "--convert-to",
            "docx",
            "--outdir",
            str(output),
            str(typed_source),
        ]
        try:
            result = subprocess.run(command, capture_output=True, check=False, timeout=limits.tool_timeout_seconds, env=env)
        except subprocess.TimeoutExpired as exc:
            raise ProcessingError("office_conversion_timeout", "Office conversion exceeded its time limit") from exc
        except OSError as exc:
            raise ProcessingError("office_runtime_unavailable", "Office conversion runtime is unavailable") from exc
        files = list(output.glob("*.docx"))
        if result.returncode != 0 or len(files) != 1:
            raise ProcessingError("office_conversion_failed", "Office conversion failed")
        if files[0].stat().st_size > limits.max_file_bytes:
            raise ProcessingError("file_size_limit", "Converted document exceeds the file size limit")
        converted = path.parent / "converted.docx"
        shutil.copyfile(files[0], converted)
        return converted


def _normalize_image_pages(path: Path, limits: ExtractionLimits) -> list[PageImage]:
    Image.MAX_IMAGE_PIXELS = limits.max_image_pixels
    pages: list[PageImage] = []
    total_bytes = 0
    try:
        with Image.open(path) as image:
            for page_number, frame in enumerate(ImageSequence.Iterator(image), start=1):
                if page_number > limits.max_pdf_pages:
                    raise ProcessingError("page_limit", "Image contains too many frames")
                width, height = frame.size
                pixels = width * height
                total_bytes += pixels * 3
                if pixels > limits.max_image_pixels or total_bytes > limits.max_decompressed_bytes:
                    raise ProcessingError("decompressed_size_limit", "Decoded image exceeds its safety limit")
                normalized = ImageOps.exif_transpose(frame.copy()).convert("RGB")
                width, height = normalized.size
                output = path.parent / f"image-page-{page_number}.png"
                normalized.save(output, format="PNG", optimize=False)
                pages.append(PageImage(page_number, output, width, height))
    except ProcessingError:
        raise
    except (OSError, UnidentifiedImageError, Image.DecompressionBombError) as exc:
        raise ProcessingError("corrupt_document", "Image is corrupt or unsupported") from exc
    return pages


def _render_pdf_page(path: Path, page: int, limits: ExtractionLimits) -> PageImage:
    prefix = path.parent / f"pdf-page-{page}"
    try:
        subprocess.run(
            [
                "pdftoppm",
                "-f",
                str(page),
                "-l",
                str(page),
                "-singlefile",
                "-r",
                "200",
                "-scale-to",
                "4096",
                "-png",
                str(path),
                str(prefix),
            ],
            capture_output=True,
            check=True,
            timeout=limits.tool_timeout_seconds,
        )
    except subprocess.TimeoutExpired as exc:
        raise ProcessingError("pdf_render_timeout", "PDF rendering exceeded its time limit", retryable=True) from exc
    except (OSError, subprocess.CalledProcessError) as exc:
        raise ProcessingError("pdf_render_failed", "PDF page rendering failed") from exc
    output = prefix.with_suffix(".png")
    try:
        with Image.open(output) as image:
            width, height = image.size
    except (OSError, UnidentifiedImageError) as exc:
        raise ProcessingError("pdf_render_failed", "Rendered PDF page is invalid") from exc
    if width * height > limits.max_image_pixels:
        raise ProcessingError("decompressed_size_limit", "Rendered PDF page exceeds its safety limit")
    return PageImage(page, output, width, height)


def _parse_pdf(
    path: Path,
    adapter: OcrAdapter,
    limits: ExtractionLimits,
    second_pass_audit: dict[str, Any],
) -> dict[str, Any]:
    try:
        reader = PdfReader(path, strict=True)
    except (OSError, PdfReadError, ValueError) as exc:
        raise ProcessingError("corrupt_document", "PDF is corrupt") from exc
    if reader.is_encrypted:
        raise ProcessingError("encrypted_document", "Encrypted documents are not supported")
    page_count = len(reader.pages)
    if not 1 <= page_count <= limits.max_pdf_pages:
        raise ProcessingError("page_limit", "PDF page count is outside the allowed range")
    native_text: dict[int, str] = {}
    missing: list[int] = []
    blocks: list[dict[str, Any]] = []
    total_chars = 0
    for page_number, page in enumerate(reader.pages, start=1):
        try:
            text = (page.extract_text() or "").strip()
        except Exception:
            text = ""
        total_chars += len(text)
        if total_chars > limits.max_text_chars:
            raise ProcessingError("text_length_limit", "Extracted text exceeds the text limit")
        native_text[page_number] = text
        if not text:
            missing.append(page_number)

    # The correction and the RECORD of it, together. Before #20 this was an in-place overwrite:
    # the pre-correction text existed nowhere afterwards and nothing said the correction had run,
    # so the one place this system rewrites what a document said was also the one place it could
    # not prove it had. `_normalize_pdf_text_layer` returns both, and the record travels in the
    # payload as evidence -- never through the sanitizer, which 0182 keeps out of this path.
    native_text, text_layer_normalizations = _normalize_pdf_text_layer(native_text)

    for page_number in sorted(native_text):
        if native_text[page_number]:
            blocks.extend(_text_blocks([native_text[page_number]], page_number, "pdf"))

    ocr_payloads: list[dict[str, Any]] = []
    # Pages nothing in this function will look at. Assigned in every branch below so the coverage
    # claim at the end reads one variable instead of re-deriving it from the output.
    unattempted: list[int] = []
    if missing and not isinstance(adapter, DisabledOcrAdapter):
        # Cap the paid path before rendering, so an oversized scan costs neither renders nor calls.
        ocr_pages = missing[: limits.max_ai_pages]
        unattempted = missing[limits.max_ai_pages :]
        # The adapter sees one memory-bounded batch at a time, so only here is the document's OCR
        # page count known. Without this the progress counter reported the batch size and restarted
        # at every flush.
        begin_progress = getattr(adapter, "begin_progress", None)
        if callable(begin_progress):
            begin_progress(len(ocr_pages))
        rendered: list[PageImage] = []
        decoded_bytes = 0

        def flush_rendered() -> None:
            nonlocal decoded_bytes
            if not rendered:
                return
            try:
                extracted = validate_extraction(adapter.extract(rendered, limits), limits)
                # Inside the try, before the rendered pages are unlinked: the recovery pass
                # re-binarizes these exact files rather than paying to render them again.
                ocr_payloads.append(
                    improve_failed_pages(
                        extracted, rendered, adapter, limits, second_pass_audit
                    )
                )
            finally:
                for rendered_page in rendered:
                    rendered_page.path.unlink(missing_ok=True)
                rendered.clear()
                decoded_bytes = 0

        try:
            for page in ocr_pages:
                rendered_page = _render_pdf_page(path, page, limits)
                page_bytes = rendered_page.width * rendered_page.height * 3
                if page_bytes > limits.max_decompressed_bytes:
                    rendered_page.path.unlink(missing_ok=True)
                    raise ProcessingError(
                        "decompressed_size_limit", "Rendered PDF page exceeds its safety limit"
                    )
                if rendered and decoded_bytes + page_bytes > limits.max_decompressed_bytes:
                    flush_rendered()
                rendered.append(rendered_page)
                decoded_bytes += page_bytes
            flush_rendered()
        finally:
            for rendered_page in rendered:
                rendered_page.path.unlink(missing_ok=True)

        for payload in ocr_payloads:
            blocks.extend(payload["blocks"])
    elif missing and len(missing) == page_count:
        raise ProcessingError("ocr_model_not_selected", "Scanned PDF requires a benchmark-approved OCR model")
    else:
        # Either every page carried a text layer -- `missing` is empty and so is this -- or OCR is
        # switched off while at least one page did carry one. That second case is not fatal by
        # design, but it does mean the image-only pages are read by nothing at all.
        unattempted = missing

    ocr_page_text: dict[int, list[str]] = {}
    if ocr_payloads:
        for payload in ocr_payloads:
            for block in payload["blocks"]:
                if block["text"]:
                    ocr_page_text.setdefault(block["page"], []).append(block["text"])
    page_text = [native_text[page] or "\n".join(ocr_page_text.get(page, [])) for page in range(1, page_count + 1)]
    plain_text = "\n\n".join(page_text)
    return _contract(
        page_count,
        plain_text,
        blocks=blocks,
        tables=[table for payload in ocr_payloads for table in payload["tables"]],
        marks=[mark for payload in ocr_payloads for mark in payload["marks"]],
        # A coverage claim and only that: these are pages this run never rendered and never sent
        # anywhere -- dropped by the `missing[: limits.max_ai_pages]` cap, or left behind because
        # no OCR adapter is configured. Nothing in this system knows what is on them.
        #
        # Deliberately NOT "a page that produced no text". A page that WAS sent and came back
        # empty is ambiguous: a blank verso and a failed read are indistinguishable without pixel
        # analysis, and inventing that heuristic is not something this parser should do. Counting
        # it would mark most scanned packets partial for having a blank back side, which is the
        # always-true flag this replaced wearing a smaller blast radius. That case is covered on
        # its own path -- `second_pass.improve` retries a page that produced zero lines.
        #
        # A PDF whose every page carried a text layer never enters the OCR branch at all and comes
        # out False, which is the whole point: it was read in full.
        partial=bool(unattempted),
        # Always present, applied or not. The PDF text layer is the only surface in this worker
        # that gets corrected, so this is the only parser that can honestly claim a decision was
        # made -- and it makes the claim on every PDF, including the ones it left alone.
        normalizations=text_layer_normalizations,
    )


def _parse_image(
    path: Path,
    adapter: OcrAdapter,
    limits: ExtractionLimits,
    second_pass_audit: dict[str, Any],
) -> dict[str, Any]:
    pages = _normalize_image_pages(path, limits)
    payload = validate_extraction(adapter.extract(pages, limits), limits)
    if payload["document"]["page_count"] != len(pages):
        raise ProcessingError("invalid_extraction", "OCR page count does not match the source")
    return improve_failed_pages(payload, pages, adapter, limits, second_pass_audit)


def extract_file(
    path: str | Path,
    claimed_mime: str | None,
    *,
    adapter: OcrAdapter | None = None,
    limits: ExtractionLimits = DEFAULT_LIMITS,
    diagnostics: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """`diagnostics`, when supplied, is filled with worker-side evidence about HOW the extraction
    was produced. It is deliberately not part of the returned payload: that payload's shape is a
    cross-surface contract validated key-for-key here and again in the Edge function, so the
    worker reports this through `resource_metadata` instead.
    """
    source = Path(path).resolve()
    detected = sniff_mime(source, claimed_mime, limits)
    if detected not in IMAGE_MIME | {
        "application/pdf",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/csv",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/rtf",
        "text/rtf",
        "text/plain",
        "text/html",
        "application/vnd.oasis.opendocument.text",
    }:
        raise ProcessingError("unsupported_mime", "Source type is not supported")
    if not mime_matches(detected, claimed_mime):
        raise ProcessingError("mime_mismatch", "Source bytes do not match the declared MIME type")
    ocr = adapter or DisabledOcrAdapter()
    second_pass_audit: dict[str, Any] = {}

    if detected in IMAGE_MIME:
        payload = _parse_image(source, ocr, limits, second_pass_audit)
    elif detected == "application/pdf":
        payload = _parse_pdf(source, ocr, limits, second_pass_audit)
    elif detected == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        payload = _parse_xlsx(source, limits)
    elif detected == "application/vnd.ms-excel":
        payload = _parse_xls(source, limits)
    elif detected == "text/csv":
        payload = _parse_csv(source, limits)
    elif detected == "text/plain":
        text = _read_utf8(source, limits)
        payload = _contract(1, text, blocks=_text_blocks([text]))
    elif detected == "text/html":
        # UNREACHABLE FROM AN UPLOAD since migration 0288 (owner ruling, OPEN-DECISIONS #346:
        # HTML is not a document type). The two client allowlists, the `documents` bucket's
        # allowed_mime_types and `public.smart_document_mime_allowed` all refuse text/html, so
        # no job the gateway can lease will ever carry it. The parser stays because deleting a
        # working extractor to enforce an intake policy is the wrong lever: if the ruling is ever
        # revisited, the type comes back by adding it to those allowlists and nothing else. It is
        # still exercised directly by self_check.py and test_scanning.py.
        payload = _parse_html(source, limits)
    elif detected == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        payload = _parse_docx(source, limits)
    elif detected in {"application/msword", "application/rtf", "text/rtf", "application/vnd.oasis.opendocument.text"}:
        suffix = {
            "application/msword": ".doc",
            "application/rtf": ".rtf",
            "text/rtf": ".rtf",
            "application/vnd.oasis.opendocument.text": ".odt",
        }[detected]
        payload = _parse_docx(_convert_to_docx(source, limits, suffix), limits)
    else:
        raise ProcessingError("unsupported_mime", "Source type is not supported")
    if diagnostics is not None and second_pass_audit:
        diagnostics["second_pass"] = second_pass_audit
    return validate_extraction(payload, limits)
